from urllib import request as rq
from urllib import parse as ps
from bs4 import BeautifulSoup as bs
import openpyxl as xl
import util
import saveToExcel as s2e

url = 'http://www.weather.go.kr/weather/forecast/mid-term-rss3.jsp'

# step 01 : parameter setting
locationDic = {'stnId':'133'}
params = ps.urlencode(locationDic)
print(params)  # stnId=133

# url 재정의
url = url + '?' + params
# print(url)

# step 02 : data load
urlObj = rq.urlopen(url)
# print(urlObj)

# step 03 : text read
urlData = urlObj.read()
# print(urlData)

# step 04 : data decode
urlText = urlData.decode('UTF-8')
# print(urlText)


html='''
<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <title>tempHTML document</title>
</head>

<body>

	<div class="top"> 

		<div>top</div>

	</div>

	<div class="nav"> 

		<div class="menu_wrap">
			<a href="#none">menu01</a> &nbsp;&nbsp;
			<a href="#none">menu02</a> &nbsp;&nbsp;
			<a href="#none">menu03</a> &nbsp;&nbsp;
		</div>


	</div>

	<div class="section"> 

		<div class="section_wrap">
			<h1>NEWS DESK</h1>
			<p id="news_desk" class="news">9시 뉴스입니다!</p>

			<h2>오늘 주요 뉴스</h2>
			<p id="today_news" class="news">오늘 뉴스입니다. 드디어 코로나19가 종식되었습니다. 마스크 벗으세요 !!</p>

			<h2>오늘 날씨</h2>
			<p id="today_weather" class="news">간절기에 접어들며 일교차가 큽니다. 감기 조심하세요</p>

			<h2>스포츠 뉴스</h2>
			<p id="sports_news" class="news">한국 여자 배구 올림픽 4강 진출</p>

		</div>


	</div>

	<div class="bottom"> 

		<div>CopyRigh</div>

	</div>

</body>
</html>
'''

# 데이터 분석 (BeautifulSoup)
soup = bs(html, 'html.parser')
# print(soup)

# ======== html hyerachy (html 계층구조 찾기) ============
h1 = soup.body.h1
print(h1)  # 테그 포함
print(h1.string)  # 테그 없이

h2 = soup.body.h2   # 다중테그일시 제일 위에있는 테그만 출력
print(h2)  # 테그 포함
print(h2.string)  # 테그 없이

# 형제 노드
h2_next = soup.body.h2.next_sibling   # (\n) 출력
print(h2_next)

h2_next_next = soup.body.h2.next_sibling.next_sibling
print(h2_next_next)

#=======================================================

# 07 : find & findall  >>>> id & class & Tag 이용
# id 찾기
news_desk = soup.find(id='news_desk')
print(news_desk)
print(news_desk.string)
# class 찾기
newsList = soup.find_all(class_='news')

for news in newsList:
    print(news)
    print(news.string)

for idx, news in enumerate(newsList):
    print(idx, news)

# a 테그 찾기
aList = soup.find_all('a')

for idx, a in enumerate(aList):
    print(a.string)
    print(a.attrs['href']) # a테그 속성값 찾기(href)

#=======================================================

# 08 select_one() & select() >>>> CSS Selector이용 >>>> 많이 쓰이고, 효율적임

news_desk = soup.select_one('#news_desk')
print(news_desk.string)

newsList = soup.select('div p.news')
for idx, news in enumerate(newsList):
    print(news)

#=======================================================

# 09 selector을 이용한 노드 선택
url = 'http://browse.gmarket.co.kr/search?keyword=%EB%82%98%EC%9D%B4%ED%82%A4%EC%9A%B4%EB%8F%99%ED%99%94'
htmlData = rq.urlopen(url)
soup = bs(htmlData, 'html.parser')
names = soup.select('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-major > div.box__item-title > span > a > span.text__item')

for idx, name in enumerate(names):
    print(f'name : {util.removeSpace(list(name)[2])}')
    # print(f'idx:{idx}, name:{name}')
    #for idx1, name1 in enumerate(names):
        # print(f'idx:{idx1}, name1:{name1}')




prices = soup.select('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-major > div.box__item-price > div > strong')
for idx, price in enumerate(prices):

     print(f'idx : {idx}, \t price : {util.wonToInt(price.string)}')


# 엑셀에 신발 이름과 가격 입력하기

# wb = xl.Workbook()
# sheet = wb.active
# sheet.title = 'ShoesInfo'
#
# col_names = ['name', 'price']
# for seq, name in enumerate(col_names):
#     sheet.cell(row=1, column=seq+1, value=name)
#
#
#
# for idx, name in enumerate(names):
#     itemData = [(list(name)[2], price.string)]
#
#
#     row_no = 2
#     for n, rows in enumerate(itemData):
#           for seq, value in enumerate(rows):
#             sheet.cell(row=row_no, column=seq+1, value=value)
#             sheet.append([list(name)[2], prices[idx].string])
#             wb.save("C:/chh_scraping/download/item.xlsx")
#             wb.close()

try:

    for i in range(len(names)):
        shoesName = util.removeSpace(list(names[i])[2])
        shooesPrice = util.wonToInt(prices[i].string)
        s2e.write2Excel([shoesName, shooesPrice])

except Exception as e:
    print(e)
    print('fail')

else:
    print('success')

